"""
Build the Excel dashboard workbook from the cleaned CSVs. All aggregation
(category summary, billing model mix, KPIs) is done with live Excel formulas
(AVERAGEIF/COUNTIFS/SUMIFS) referencing the raw data sheets, not pasted-in
Python-computed numbers, so the workbook recalculates if the underlying rows
change.

Run: python3 python/03_build_excel_dashboard.py
Then: python3 <xlsx-skill>/scripts/recalc.py excel/saas_market_dashboard.xlsx
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
PROC = ROOT / "data" / "processed"
OUT = ROOT / "excel" / "saas_market_dashboard.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=18, color="2F5496")
SUBTITLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="595959")
LABEL_FONT = Font(name=FONT_NAME, bold=True, size=11)
BODY_FONT = Font(name=FONT_NAME, size=10)
KPI_FONT = Font(name=FONT_NAME, bold=True, size=20, color="2F5496")
KPI_LABEL_FONT = Font(name=FONT_NAME, size=10, color="595959")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_df_as_table(ws, df, start_row, table_name):
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col)
    style_header_row(ws, start_row, len(df.columns))
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
    last_row = start_row + len(df)
    last_col = get_column_letter(len(df.columns))
    ref = f"A{start_row}:{last_col}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
    )
    ws.add_table(table)
    return last_row


def main() -> None:
    products = pd.read_csv(PROC / "products_clean.csv")
    plans = pd.read_csv(PROC / "pricing_plans_clean.csv")

    product_cols = [
        "product_id", "name", "category_slug", "category_name",
        "overall_rating", "g2_rating", "capterra_rating", "founded_year",
        "hq_location", "has_free_tier", "has_free_trial", "has_custom_pricing",
        "price_comparable", "entry_price", "max_price_monthly", "plan_count",
        "feature_count", "website_url",
    ]
    products_out = products[product_cols].copy()

    plan_cols = ["id", "product_name", "plan_name", "price_monthly", "price_annual", "billing_model"]
    plans_out = plans[plan_cols].copy()

    categories = sorted(products["category_name"].dropna().unique().tolist())

    wb = Workbook()

    # ---------------------------------------------------------------
    # Sheet: Products (raw cleaned data)
    # ---------------------------------------------------------------
    ws_p = wb.active
    ws_p.title = "Products"
    last_row_p = write_df_as_table(ws_p, products_out, 1, "ProductsTable")
    n_products = len(products_out)

    # Helper column: entry price only when price_comparable (excludes DEX's
    # swap-fee-as-price rows) so MEDIAN/AVERAGE formulas elsewhere can just
    # point at this column without re-deriving the exclusion each time.
    helper_col = len(product_cols) + 1
    helper_letter = get_column_letter(helper_col)
    ws_p.cell(row=1, column=helper_col, value="entry_price_comparable")
    ws_p.cell(row=1, column=helper_col).fill = HEADER_FILL
    ws_p.cell(row=1, column=helper_col).font = HEADER_FONT
    price_comparable_col = get_column_letter(product_cols.index("price_comparable") + 1)
    entry_price_col = get_column_letter(product_cols.index("entry_price") + 1)
    for r in range(2, last_row_p + 1):
        ws_p.cell(
            row=r, column=helper_col,
            value=f'=IF({price_comparable_col}{r},{entry_price_col}{r},"")',
        )
    ws_p.freeze_panes = "A2"
    autosize(ws_p, [10, 22, 20, 22, 10, 10, 12, 10, 24, 10, 10, 12, 12, 10, 12, 10, 10, 30, 16])

    # ---------------------------------------------------------------
    # Sheet: Pricing Plans (raw cleaned data)
    # ---------------------------------------------------------------
    ws_pl = wb.create_sheet("Pricing Plans")
    write_df_as_table(ws_pl, plans_out, 1, "PricingPlansTable")
    ws_pl.freeze_panes = "A2"
    autosize(ws_pl, [8, 24, 20, 14, 14, 14])
    n_plans = len(plans_out)

    # ---------------------------------------------------------------
    # Sheet: Category Summary (formula-driven aggregation)
    # ---------------------------------------------------------------
    ws_c = wb.create_sheet("Category Summary")
    headers = [
        "category_name", "product_count", "avg_rating", "avg_entry_price",
        "free_tier_count", "free_tier_pct", "avg_feature_count",
    ]
    for j, h in enumerate(headers, start=1):
        ws_c.cell(row=1, column=j, value=h)
    style_header_row(ws_c, 1, len(headers))

    cat_name_col = "D"  # Products!D = category_name
    rating_col = "E"
    entry_helper_col = helper_letter
    free_tier_col = "J"
    feature_count_col = "Q"

    for i, cat in enumerate(categories, start=2):
        ws_c.cell(row=i, column=1, value=cat).font = BODY_FONT
        ws_c.cell(row=i, column=2,
            value=f'=COUNTIF(Products!{cat_name_col}2:{cat_name_col}{last_row_p},A{i})').font = BODY_FONT
        ws_c.cell(row=i, column=3,
            value=f'=ROUND(AVERAGEIF(Products!{cat_name_col}2:{cat_name_col}{last_row_p},A{i},'
                  f'Products!{rating_col}2:{rating_col}{last_row_p}),2)').font = BODY_FONT
        ws_c.cell(row=i, column=4,
            value=f'=IFERROR(ROUND(AVERAGEIF(Products!{cat_name_col}2:{cat_name_col}{last_row_p},A{i},'
                  f'Products!{entry_helper_col}2:{entry_helper_col}{last_row_p}),2),"N/A")').font = BODY_FONT
        ws_c.cell(row=i, column=5,
            value=f'=COUNTIFS(Products!{cat_name_col}2:{cat_name_col}{last_row_p},A{i},'
                  f'Products!{free_tier_col}2:{free_tier_col}{last_row_p},TRUE)').font = BODY_FONT
        ws_c.cell(row=i, column=6, value=f'=ROUND(E{i}/B{i}*100,1)').font = BODY_FONT
        ws_c.cell(row=i, column=7,
            value=f'=ROUND(AVERAGEIF(Products!{cat_name_col}2:{cat_name_col}{last_row_p},A{i},'
                  f'Products!{feature_count_col}2:{feature_count_col}{last_row_p}),2)').font = BODY_FONT

    last_row_c = len(categories) + 1
    ref = f"A1:G{last_row_c}"
    table = Table(displayName="CategorySummaryTable", ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws_c.add_table(table)
    ws_c.freeze_panes = "A2"
    autosize(ws_c, [22, 14, 12, 16, 14, 14, 16])

    note_row = last_row_c + 2
    ws_c.cell(row=note_row, column=1,
        value="Note: 'DEX' avg_entry_price shows N/A — its listed prices are swap-fee %, not USD/month subscriptions.").font = SUBTITLE_FONT

    # ---------------------------------------------------------------
    # Sheet: Billing Models (formula-driven)
    # ---------------------------------------------------------------
    ws_b = wb.create_sheet("Billing Models")
    ws_b.cell(row=1, column=1, value="billing_model")
    ws_b.cell(row=1, column=2, value="plan_count")
    ws_b.cell(row=1, column=3, value="pct_of_plans")
    style_header_row(ws_b, 1, 3)
    billing_models = ["flat", "per-seat", "usage-based", "per-project"]
    billing_col = "F"  # Pricing Plans!F = billing_model
    for i, bm in enumerate(billing_models, start=2):
        ws_b.cell(row=i, column=1, value=bm).font = BODY_FONT
        ws_b.cell(row=i, column=2,
            value=f"=COUNTIF('Pricing Plans'!{billing_col}2:{billing_col}{n_plans + 1},A{i})").font = BODY_FONT
        ws_b.cell(row=i, column=3, value=f"=ROUND(B{i}/SUM($B$2:$B$5)*100,1)").font = BODY_FONT
    last_row_b = len(billing_models) + 1
    table_b = Table(displayName="BillingModelTable", ref=f"A1:C{last_row_b}")
    table_b.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws_b.add_table(table_b)
    autosize(ws_b, [16, 14, 14])

    # ---------------------------------------------------------------
    # Sheet: Overview (KPIs + methodology notes)
    # ---------------------------------------------------------------
    ws_o = wb.create_sheet("Overview")
    ws_o.sheet_view.showGridLines = False
    ws_o.column_dimensions["A"].width = 4
    for col, w in zip("BCDEFG", [24, 24, 24, 24, 24, 24]):
        ws_o.column_dimensions[col].width = w

    ws_o["B2"] = "SaaS Market 2026 — Pricing & Positioning Analysis"
    ws_o["B2"].font = TITLE_FONT
    ws_o["B3"] = "Source: ComparEdge (Kaggle) — 331 SaaS products across 28 categories, snapshot dated 2026-04-27"
    ws_o["B3"].font = SUBTITLE_FONT

    kpis = [
        ("Total Products", f"=COUNTA(Products!B2:B{last_row_p})"),
        ("Categories", f"=COUNTA('Category Summary'!A2:A{last_row_c})"),
        ("Median Entry Price", f'=TEXT(MEDIAN(Products!{helper_letter}2:{helper_letter}{last_row_p}),"$0.00/mo")'),
        ("% With Free Tier", f'=ROUND(COUNTIF(Products!J2:J{last_row_p},TRUE)/COUNTA(Products!B2:B{last_row_p})*100,1)&"%"'),
    ]

    start_col = 2
    for i, (label, formula) in enumerate(kpis):
        col_letter = get_column_letter(start_col + i)
        ws_o[f"{col_letter}5"] = label
        ws_o[f"{col_letter}5"].font = KPI_LABEL_FONT
        ws_o[f"{col_letter}6"] = formula
        ws_o[f"{col_letter}6"].font = KPI_FONT

    ws_o["B9"] = "How to read this workbook"
    ws_o["B9"].font = LABEL_FONT
    notes = [
        "Products — one row per SaaS product (cleaned from a SQLite relational export). "
        "'entry_price_comparable' is a helper column used by every downstream formula.",
        "Pricing Plans — one row per pricing tier per product (1,013 total).",
        "Category Summary — category-level stats computed live with AVERAGEIF/COUNTIFS "
        "formulas against the Products sheet, not pasted-in numbers.",
        "Billing Models — mix of flat / per-seat / usage-based / per-project billing "
        "across all pricing plans.",
        "Dashboard — charts built from the Category Summary and Billing Models tables.",
    ]
    for i, note in enumerate(notes, start=10):
        ws_o[f"B{i}"] = f"• {note}"
        ws_o[f"B{i}"].font = BODY_FONT
        ws_o.merge_cells(f"B{i}:G{i}")
        ws_o[f"B{i}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_o.row_dimensions[i].height = 28

    dq_row = 17
    ws_o[f"B{dq_row}"] = "Known data quality issues (see sql/02_data_quality_checks.sql for detection queries)"
    ws_o[f"B{dq_row}"].font = LABEL_FONT
    dq_notes = [
        "'ai-tools' was a phantom umbrella category (104 products claimed, 0 actually "
        "tagged) — removed; the real AI products live in the ai-* subcategories + LLM.",
        "'DEX' pricing_plans.price_monthly stores swap-fee percentages, not USD/month "
        "subscription prices — excluded from all price comparisons (price_comparable = FALSE).",
        "products.review_count is 0 for every row (unpopulated in the source data) — not "
        "used anywhere as a reliability signal.",
        "~11 products have no overall_rating and ~89 have at least one custom/'Contact "
        "Sales' plan — both kept as-is (blank/N/A) rather than guessed at.",
        "A few category names were auto-titlecased acronyms ('Crm', 'Dex', 'Vpn') — "
        "corrected to CRM, DEX, VPN, DeFi Tools.",
    ]
    for i, note in enumerate(dq_notes, start=dq_row + 1):
        ws_o[f"B{i}"] = f"• {note}"
        ws_o[f"B{i}"].font = BODY_FONT
        ws_o.merge_cells(f"B{i}:G{i}")
        ws_o[f"B{i}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_o.row_dimensions[i].height = 28

    # ---------------------------------------------------------------
    # Sheet: Dashboard (charts)
    # ---------------------------------------------------------------
    ws_d = wb.create_sheet("Dashboard")
    ws_d.sheet_view.showGridLines = False
    ws_d["B2"] = "SaaS Market 2026 — Dashboard"
    ws_d["B2"].font = TITLE_FONT

    # Chart 1: Avg entry price by category
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.title = "Average Entry Price by Category (USD/month)"
    chart1.y_axis.title = None
    chart1.x_axis.title = "USD / month"
    data1 = Reference(ws_c, min_col=4, min_row=1, max_row=last_row_c)
    cats1 = Reference(ws_c, min_col=1, min_row=2, max_row=last_row_c)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.height, chart1.width = 10, 18
    chart1.legend = None
    ws_d.add_chart(chart1, "B4")

    # Chart 2: Free tier % by category
    chart2 = BarChart()
    chart2.type = "bar"
    chart2.title = "Free-Tier Adoption Rate by Category (%)"
    data2 = Reference(ws_c, min_col=6, min_row=1, max_row=last_row_c)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats1)
    chart2.height, chart2.width = 10, 18
    chart2.legend = None
    ws_d.add_chart(chart2, "L4")

    # Chart 3: Billing model mix (pie)
    chart3 = PieChart()
    chart3.title = "Billing Model Mix (share of pricing plans)"
    data3 = Reference(ws_b, min_col=2, min_row=1, max_row=last_row_b)
    cats3 = Reference(ws_b, min_col=1, min_row=2, max_row=last_row_b)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.height, chart3.width = 10, 14
    ws_d.add_chart(chart3, "B22")

    # Chart 4: Avg rating by category
    chart4 = BarChart()
    chart4.type = "bar"
    chart4.title = "Average Rating by Category"
    data4 = Reference(ws_c, min_col=3, min_row=1, max_row=last_row_c)
    chart4.add_data(data4, titles_from_data=True)
    chart4.set_categories(cats1)
    chart4.height, chart4.width = 10, 14
    chart4.legend = None
    ws_d.add_chart(chart4, "L22")

    desired_order = [
        "Overview", "Dashboard", "Products", "Pricing Plans",
        "Category Summary", "Billing Models",
    ]
    wb._sheets = [wb[name] for name in desired_order]
    wb.active = 0

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved {OUT} ({n_products} products, {n_plans} plans, {len(categories)} categories)")


if __name__ == "__main__":
    main()
